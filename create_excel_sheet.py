#!/usr/bin/env python3
"""
Script to create TestStub.xlsx with login credentials
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "LoginData"

# Add headers (first row)
headers = ["TestCaseID", "Username", "Password", "ExpectedResult"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add login test data
test_data = [
    ["TC_001", "Admin", "admin123", "Login Successful"],
    ["TC_002", "employee1", "emp123456", "Login Successful"],
    ["TC_003", "manager1", "mgr123456", "Login Successful"],
    ["TC_004", "invaliduser", "wrongpass", "Login Failed"],
]

for row_num, row_data in enumerate(test_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20

# Save the workbook
output_path = r"src\test\resources\TestStub.xlsx"
wb.save(output_path)
print(f"✓ Excel file created successfully at: {output_path}")
print("\nLogin Credentials added:")
print("=" * 70)
for row_data in test_data:
    print(f"  TestID: {row_data[0]} | Username: {row_data[1]} | Password: {row_data[2]}")
print("=" * 70)
